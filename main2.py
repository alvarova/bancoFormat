#!/usr/bin/env python
# -*- coding: utf-8 -*-

def preRunning(list):
    error=0
    print("Procesando:")
    for libreria in list:
        print("-"+libreria)
        try:
            __import__(libreria)
        except ImportError:
            print("  Falta la librería "+libreria+". Instala con: pip install "+libreria)
            error+=1
    if error>0:
        print("Verificar. Falta/n "+str(error)+" librerias que son requeridas para ejecutar la app.")
        print("Saliendo de la aplicación.")
        exit()

librerias= ['kivy','openpyxl','kivymd']


preRunning(librerias)

import os
#from os import startfile
import subprocess
import time
import openpyxl
from datetime import datetime


from kivy.core.window import Window
from kivy.lang import Builder

from kivy.uix.screenmanager import ScreenManager, Screen

from kivymd.uix.menu import MDDropdownMenu
from kivymd.app import MDApp
from kivymd.uix.filemanager import MDFileManager
from kivymd.toast import toast
from kivy.properties import BooleanProperty
from kivy.properties import StringProperty


class Ui(ScreenManager):
    pass

class Aplicacion(MDApp):

    processFile = StringProperty("")
    is_process_file_empty = BooleanProperty(True) 
    
    def events(self, window, key, scancode, codepoint, modifier):
        # Aquí manejas el evento del teclado como prefieras
        print(f"Tecla presionada: {key}")
        if key == 27:  # Esto es un ejemplo, 27 es el código de la tecla 'ESC'
            Aplicacion.get_running_app().stop()
            return True  # Devuelve True para indicar que has manejado el evento
        return False  # Devuelve False para que otros manejadores puedan procesarlo

    def __init__(self, **kwargs):
        super().__init__(**kwargs)        
        Window.bind(on_keyboard=self.events)
        self.manager_open = False
        self.file_manager = MDFileManager(
            exit_manager=self.exit_manager, select_path=self.select_path
        )

    def on_start(self):
        menu_items = [
                    {
                        "text": "Cargar planilla",
                        "icon": "file-upload",
                        "on_release": lambda x="Cargar planilla": self.menu_callback(x),
                    },
                    {
                        "text": "Mostrar planilla",
                        "icon": "file-document-outline",
                        "on_release": lambda x="Mostrar planilla": self.menu_callback(x),
                    },
                    {
                        "text": "Procesar planilla",
                        "icon": "file-check",
                        "on_release": lambda x="Procesar planilla": self.menu_callback(x),
                    },
                ]        
        # Asegúrate de que el acceso a self.root.ids.menu_button se realice después de que la interfaz gráfica se haya cargado completamente
        menu_button = self.root.ids.menu_button
        print("Menu button loaded:", menu_button)
        # Inicializa el menú
        self.menu = MDDropdownMenu(
            caller=self.root.ids.menu_button,
            items=menu_items,
            width_mult=4,
            position='top',
        )
    def menu_callback(self, text_item):
        if text_item == "Cargar planilla":
            self.file_manager_open()
        elif text_item == "Mostrar planilla":
            self.abrirArchivo()
        elif text_item == "Procesar planilla":
            self.procesa()
    def open_menu(self, *args):
        # Método para abrir el menú
        self.menu.open()

    def valida_digito(self, suc, cta):
        """Valida el dígito verificador de una cuenta bancaria.

        Args:
            suc: Número de sucursal (cadena de 3 dígitos).
            cta: Número de cuenta (cadena de 7 dígitos).

        Returns:
            True si el dígito verificador es correcto, False en caso contrario.
        """
        print("===> In Value Suc/Cta:"+str(suc)+"/"+str(cta))
        def nulo_cero(val):
            return val if val is not None else 0

        cdig = str(cta)[-2:]
        idig = int(cdig)
        print("===> In Value cDig:"+str(cdig))
        suc=suc.zfill(3)   # Formatea sucursal a 3 dígitos
        suc=suc[-3:]
        cta = str(cta).zfill(9)  # Rellenar con ceros a la izquierda hasta 9 dígitos
        cta = cta[:7]          # Tomar los primeros 7 caracteres


        v1 = int(suc[0]) * 9
        v2 = int(suc[1]) * 8
        v3 = int(suc[2]) * 7

        v4 = int(cta[0]) * 8
        v5 = int(cta[1]) * 9
        v6 = int(cta[2]) * 6
        v7 = int(cta[3]) * 5
        v8 = int(cta[4]) * 4
        v9 = int(cta[5]) * 3
        v10 = int(cta[6]) * 2
        print("+SUC:"+suc)
        print("+CTA:"+cta)
        
        vdigito = v1 + v2 + v3 + v4 + v5 + v6 + v7 + v8 + v9 + v10
        
        print("++++"+str(vdigito))
        
        vdigito = 11 - (vdigito % 11)
        
        if vdigito == 11:
            vdigito = 0
        vale = str(vdigito == idig)
        print ("Valida:"+vale)
        return vdigito == idig  
      
    def build(self):
        self.icon="icon.png"
        self.theme_cls.primary_palette = 'Green'
        #self.theme_cls.theme_style = 'Dark'
        self.root = Builder.load_file('ui.kv')
        return self.root
  
    def file_manager_open(self):
        self.file_manager.show(os.path.expanduser("~"))  # output manager to the screen
        self.manager_open = True
    
    def select_path(self, path: str):
        '''
        It will be called when you click on the file name
        or the catalog selection button.

        :param path: path to the selected directory or file;
        '''

        self.exit_manager()
        toast(path)
        #self.processFile = path
        self.set_process_file(path)


    #def exit_manager(self, *args):
        '''Called when the user reaches the root of the directory tree.'''

    #    self.manager_open = False
    #    self.file_manager.close()        
    def exit_manager(self, *args):
        if self.file_manager._window_manager and self.file_manager._window_manager.parent:
            self.file_manager.close()
        else:
            print("No se puede cerrar, ya ha sido cerrado o nunca fue inicializado.")
            
    def abrirArchivo(self): 
        # startfile(self.processFile)  windows
        subprocess.run(["xdg-open", self.processFile])

    def update_button_state(self):
        self.is_process_file_empty = not bool(self.processFile)

    def set_process_file(self, file_path):
        self.processFile = file_path
        self.update_button_state()


    def cargar_excel(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        return workbook.active

    def determinar_tipo_archivo(self, sheet):
        datoCelda = sheet.cell(row=1, column=6)
        if datoCelda.value is None:
            return 0  # Planilla original
        else:
            return 1  # Planilla nueva

    def obtener_parametros_fijos(self, sheet, addColumna):
        empresa = str(sheet.cell(row=1, column=5+addColumna).value).zfill(4)
        convenio = str(sheet.cell(row=2, column=5+addColumna).value).zfill(4)
        fecha_envio = str(sheet.cell(row=3, column=5+addColumna).value).replace("-", "").split(" ")[0]
        return empresa, convenio, fecha_envio

    def calcular_importe_total(self, sheet, addColumna):
        importe_total = 0
        for row in range(7, sheet.max_row + 1):
            importe = sheet.cell(row=row, column=5+addColumna).value
            if importe is not None:
                importe_total += int(importe * 1000)
        return str(importe_total).zfill(14)

    def generar_cabecera(self, empresa, convenio, fecha_envio, importe_total):
        cabecera = f"9998{empresa}{convenio}{fecha_envio[0:4]}{fecha_envio[4:6]}{fecha_envio[6:8]}000001{importe_total}"
        cabecera += " " * 79 + "\r\n"
        return cabecera

    def generar_lotes(self, sheet, addColumna, empresa, convenio, fecha_envio, valida_digito):
        lotes = []
        lotesNo = []
        row = 7
        cuenta = 0

        while True:
            sistema = str(sheet.cell(row=row, column=1+addColumna).value).zfill(2)
            sucursal = str(sheet.cell(row=row, column=2+addColumna).value).zfill(4)
            nro_cuenta = str(sheet.cell(row=row, column=3+addColumna).value).zfill(12)
            validaCuenta = str(sheet.cell(row=row, column=3+addColumna).value).zfill(9)
            nro_comprobante = str(sheet.cell(row=row, column=4+addColumna).value).zfill(10)
            importe = str(sheet.cell(row=row, column=5+addColumna).value * 1000).zfill(14)
            cuenta += 1

            if valida_digito(sucursal, validaCuenta):
                lote = f"{empresa}{sistema}{sucursal}{nro_cuenta}02{importe}{fecha_envio[0:4]}{fecha_envio[4:6]}{fecha_envio[6:8]}00{nro_comprobante}{convenio}9999"
                lote += " " * 40 + "0000000000000" + "\r\n"
                lotes.append(lote)
            else:
                lotesNo.append(nro_cuenta + "\r\n")

            row += 1
            if sheet.cell(row=row, column=1).value is None:
                break

        return lotes, lotesNo, cuenta

    def guardar_archivos(self, processFile, cabecera, lotes, lotesNo):
        now = datetime.now()
        cadena = now.strftime("%Y%m%d%H%M%S")
        with open(processFile + "-" + cadena + ".txt", "w") as archivo:
            archivo.write(cabecera)
            for lote in lotes:
                archivo.write(lote)

        if len(lotesNo) != 0:
            cabecera_error = "Las siguientes cuentas no coinciden con el numero de sucursal. Revisar nuevamente. Estas cuentas se omitieron.\r\n"
            with open("ERROR-" + cadena + ".txt", "w") as archivoN:
                archivoN.write(cabecera_error)
                for loteN in lotesNo:
                    archivoN.write(loteN)

    def procesa(self):
        if len(self.processFile) != 0:
            print('iniciando proceso con ' + self.processFile)
            self.root.ids.box.children[1].text = 'Procesando: ' + self.processFile

            sheet = self.cargar_excel(self.processFile)
            addColumna = self.determinar_tipo_archivo(sheet)
            empresa, convenio, fecha_envio = self.obtener_parametros_fijos(sheet, addColumna)
            importe_total = self.calcular_importe_total(sheet, addColumna)
            cabecera = self.generar_cabecera(empresa, convenio, fecha_envio, importe_total)
            lotes, lotesNo, cuenta = self.generar_lotes(sheet, addColumna, empresa, convenio, fecha_envio, self.valida_digito)
            self.guardar_archivos(self.processFile, cabecera, lotes, lotesNo)
            # Actualizar el estado del Label
            self.root.ids.estado_label.text = f"Resultado: {len(lotes)} de {cuenta}"
            
            if len(lotes) < cuenta:
                self.root.ids.estado_label.color = (0.86, 0.08, 0.24, 1)  # Rojo (RGBA)
            else:
                self.root.ids.estado_label.color = (0.13, 0.55, 0.13, 1)  # Verde (RGBA)

            toast(self.processFile + "-" + ".txt' generado.")
            self.root.ids.box.children[1].text = 'Procesado: ' + self.processFile
            print(f"Total de registros guardados {len(lotes)} de {cuenta}")
        else:
            toast("Para continuar, debe seleccione la planilla a procesar")
            


Aplicacion().run()